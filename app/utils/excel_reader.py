from pathlib import Path

from openpyxl import load_workbook


class ExcelReader:
    """
    Reads email records from an Excel (.xlsx) file.
    """

    REQUIRED_COLUMNS = [
        "Sample#",
        "Email Sample",
        "Category",
        "Classification",
    ]

    def read(self, excel_path: Path) -> list[dict]:
        """
        Read Excel file and return list of rows.

        Args:
            excel_path: Path to Excel file.

        Returns:
            List of dictionaries.
        """

        workbook = load_workbook(
            filename=excel_path,
            data_only=True
        )

        sheet = workbook.active

        rows = list(sheet.iter_rows(values_only=True))

        if not rows:
            raise ValueError("Excel file is empty.")

        headers = [str(col).strip() if col else "" for col in rows[0]]

        self._validate_headers(headers)

        records = []

        for row in rows[1:]:

            # Skip completely empty rows
            if all(cell is None for cell in row):
                continue

            record = dict(zip(headers, row))

            records.append(
                {
                    "sample_id": record.get("Sample#"),
                    "email_text": record.get("Email Sample"),
                    "category": record.get("Category"),
                    "classification": record.get("Classification"),
                }
            )

        workbook.close()

        return records

    def _validate_headers(self, headers: list[str]) -> None:
        """
        Validate required columns exist.
        """

        missing = [
            column
            for column in self.REQUIRED_COLUMNS
            if column not in headers
        ]

        if missing:
            raise ValueError(
                f"Missing required columns: {', '.join(missing)}"
            )

        